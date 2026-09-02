#!/usr/bin/env python3
"""Convert every sheet of an .xlsx workbook to JSON: {sheetName: [{col: val, ...}, ...]}.

Exists because exceljs (this project's usual .xlsx reader, see
scripts/import-exercises.ts) cannot parse some real-world workbooks — found
on the 2026-09-02 fresh site extracts in data/source/, whose workbook.xml
uses a namespace-prefixed root element (<x:workbook xmlns:x="...">) instead
of the default-namespace form real Excel emits (<workbook xmlns="...">).
Both are valid OOXML; exceljs's simplified parser only recognizes the
unprefixed form. openpyxl handles both correctly, so this script is the
Python fallback for exactly that case — used by
scripts/import-workout-extract.ts via a subprocess call, not directly.

Requires openpyxl: pip install openpyxl
"""
import json
import sys


def main():
    if len(sys.argv) != 2:
        print("Usage: xlsx_to_json.py <path-to-xlsx>", file=sys.stderr)
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print(
            "openpyxl is required but not installed. Run: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    path = sys.argv[1]
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)

    result = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, [])
        headers = [str(h).strip() if h is not None else None for h in header_row]

        sheet_rows = []
        for row in rows:
            if all(value is None for value in row):
                continue
            record = {
                header: value
                for header, value in zip(headers, row)
                if header is not None
            }
            sheet_rows.append(record)
        result[sheet_name] = sheet_rows

    json.dump(result, sys.stdout, default=str)


if __name__ == "__main__":
    main()
